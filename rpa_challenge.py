import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pyautogui as py

url = 'https://rpachallenge.com/'

path = r'D:\Workspace\RPA Challenge\challenge.xlsx'
wb = openpyxl.load_workbook(path)
sh1 = wb.active
r_count = sh1.max_row
c_count = sh1.max_column

print(f'Row count: {r_count}, Column count: {c_count}')

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
chrome_options.binary_location = './/CFT driver/chrome-win64/chrome.exe'
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), 
                          options=chrome_options)
driver.get(url)
driver.maximize_window()
driver.implicitly_wait(5)
# ------------------------------------
driver.find_element(By.XPATH, "//*[contains(@class, 'waves-effect col s12 m12 l12 btn-large uiColorButton')]").click()
for r in range (2, r_count + 1):
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelFirstName')]").send_keys(sh1.cell(row=r, column=1).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelLastName')]").send_keys(sh1.cell(row=r, column=2).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelCompanyName')]").send_keys(sh1.cell(row=r, column=3).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelRole')]").send_keys(sh1.cell(row=r, column=4).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelAddress')]").send_keys(sh1.cell(row=r, column=5).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelEmail')]").send_keys(sh1.cell(row=r, column=6).value)
    driver.find_element(By.XPATH, "//*[contains(@ng-reflect-name, 'labelPhone')]").send_keys(sh1.cell(row=r, column=7).value)
    driver.find_element(By.XPATH, "//*[contains(@value, 'Submit')]").click()
    sh1.cell(row=r, column=8).value = 'Completed'

py.alert('All records submitted successfully!')
wb.save(path)